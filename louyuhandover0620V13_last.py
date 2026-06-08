import pandas as pd
from geopy.distance import geodesic
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import messagebox
import sys

# 获取资源路径 - 用于打包后的exe正确获取文件路径dfadfasdgaf的风格
def resource_path(relative_path):
    """获取资源的绝对路径，支持开发环境和PyInstaller打包环境"""
    try:
        # PyInstaller创建临时文件夹并将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except AttributeError:
        # 开发环境下直接返回当前路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)
     # PyInstaller创建临时
     
# 计算经纬度之间的距离
def calculate_distance(row, df):
    """
    计算经纬度之间的距离
    :param row: 数据行
    :param df: 数据框
    :return: 距离（米）或 None
    """
    source_lon = row[df.columns[1]]
    source_lat = row[df.columns[2]]
    target_lon = row[df.columns[4]]
    target_lat = row[df.columns[5]]
    try:
        if pd.notna(source_lon) and pd.notna(source_lat) and pd.notna(target_lon) and pd.notna(target_lat):
            source_coords = (source_lat, source_lon)
            target_coords = (target_lat, target_lon)
            return geodesic(source_coords, target_coords).meters
        return None
    except (ValueError, TypeError):
        return None

# 计算百分比
def calculate_percentage(row, df):
    """
    计算百分比
    :param row: 数据行
    :param df: 数据框
    :return: 百分比字符串或 None
    """
    col_12 = row[df.columns[11]]
    col_13 = row[df.columns[12]]
    if pd.notna(col_12) and col_12 != 0:
        return f"{(col_13 / col_12) * 100:.2f}%"
    return None

# 判断第 15 列的值
def judge_condition(row, cishu_value, juli_value, bili_value, df):
    """
    判断第 15 列的值
    :param row: 数据行
    :param cishu_value: 最小切换次数
    :param juli_value: 超远切换距离
    :param bili_value: 低切换成功率比例
    :param df: 数据框
    :return: 判断结果或 None
    """
    col_11 = row[df.columns[10]]
    col_12 = row[df.columns[11]]
    col_14_str = row[df.columns[13]]
    if pd.notna(col_12) and col_12 > cishu_value:
        if col_14_str:
            try:
                col_14 = float(col_14_str.rstrip('%'))
                if col_14 < bili_value:
                    return "低切换成功率"
            except ValueError:
                pass
        if pd.notna(col_11) and col_11 > juli_value:
            return "远距离切换"
    return None

# 进行 Vlookup 操作
def perform_vlookup(df, df_sheet2):
    """
    从 Sheet2 中进行 Vlookup 操作，填充 Sheet1 中的经度和纬度
    :param df: Sheet1 数据框
    :param df_sheet2: Sheet2 数据框
    :return: 填充后的 Sheet1 数据框
    """
    # 创建一个字典用于 Vlookup，将 Sheet2 第 1 列和第 2 列组合成字典，用于查找经度
    lookup_dict_lon = dict(zip(df_sheet2.iloc[:, 0], df_sheet2.iloc[:, 1]))
    # 创建一个字典用于 Vlookup，将 Sheet2 第 1 列和第 3 列组合成字典，用于查找纬度
    lookup_dict_lat = dict(zip(df_sheet2.iloc[:, 0], df_sheet2.iloc[:, 2]))

    # 根据 Sheet1 第 1 列的内容从 Sheet2 中查找对应经度并填入第 2 列
    df[df.columns[1]] = df[df.columns[0]].map(lookup_dict_lon)
    # 根据 Sheet1 第 1 列的内容从 Sheet2 中查找对应纬度并填入第 3 列
    df[df.columns[2]] = df[df.columns[0]].map(lookup_dict_lat)
    # 根据 Sheet1 第 4 列的内容从 Sheet2 中查找对应经度并填入第 5 列
    df[df.columns[4]] = df[df.columns[3]].map(lookup_dict_lon)
    # 根据 Sheet1 第 4 列的内容从 Sheet2 中查找对应纬度并填入第 6 列
    df[df.columns[5]] = df[df.columns[3]].map(lookup_dict_lat)

    # 检查是否有未匹配的值
    if df[df.columns[1]].isna().any() or df[df.columns[2]].isna().any() or df[df.columns[4]].isna().any() or df[df.columns[5]].isna().any():
        messagebox.showwarning("警告", "在 Vlookup 过程中，部分数据未找到匹配项，对应单元格将填充为 NaN。")

    return df

# 数据处理函数
def process_data(df, cishu_value, juli_value, bili_value):
    """
    处理数据，包括计算距离、百分比和判断条件
    :param df: 数据框
    :param cishu_value: 最小切换次数
    :param juli_value: 超远切换距离
    :param bili_value: 低切换成功率比例
    :return: 处理后的数据框
    """
    # 从第 1 行（索引为 0）开始计算距离并填入第 11 列
    df.loc[0:, df.columns[10]] = df.loc[0:].apply(lambda row: calculate_distance(row, df), axis=1)
    # 第 8 列 + 第 10 列之和填入第 12 列
    df[df.columns[11]] = df[df.columns[7]] + df[df.columns[9]]
    # 第 7 列 + 第 9 列之和填入第 13 列
    df[df.columns[12]] = df[df.columns[6]] + df[df.columns[8]]
    # 第 13 列除以第 12 列，结果以百分比形式填入第 14 列
    df[df.columns[13]] = df.apply(lambda row: calculate_percentage(row, df), axis=1)
    # 定义第 15 列的判断逻辑
    df[df.columns[14]] = df.apply(lambda row: judge_condition(row, cishu_value, juli_value, bili_value, df), axis=1)
    return df

# 统计结果
def count_results(df):
    """
    统计第 15 列中“低切换成功率”和“远距离切换”的数量
    :param df: 数据框
    :return: 低切换成功率数量和远距离切换数量
    """
    low_success_count = df[df.columns[14]].value_counts().get("低切换成功率", 0)
    long_distance_count = df[df.columns[14]].value_counts().get("远距离切换", 0)
    return low_success_count, long_distance_count

# 保存数据并填充颜色
def save_and_fill(df, file_path):
    """
    将处理后的数据保存到文件，对特定单元格填充颜色，同时保留 Sheet2
    :param df: 数据框
    :param file_path: 文件路径
    """
    # 加载工作簿
    wb = load_workbook(file_path)
    # 获取 Sheet1 工作表
    ws = wb['Sheet1']
    # 清空 Sheet1 中的原有数据（除表头）
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    # 将处理后的数据写入 Sheet1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # 填充颜色
    col_15_index = df.columns.get_loc(df.columns[14]) + 1
    for row in ws.iter_rows(min_row=2, min_col=col_15_index, max_col=col_15_index):
        cell = row[0]
        if cell.value == "远距离切换":
            cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        elif cell.value == "低切换成功率":
            cell.fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
    # 保存工作簿
    wb.save(file_path)

# 使用文件选择对话框让用户选择Excel文件
def select_file():
    from tkinter import filedialog
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)
        check_inputs()  # 检查输入是否完整

# 运行主程序
def run_program():
    try:
        # 获取输入框的值
        cishu_value = int(cishu_entry.get())
        juli_value = int(juli_entry.get())
        bili_value = int(bili_entry.get())
        file_path = file_entry.get()
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            messagebox.showerror("错误", f"文件 {file_path} 不存在！\n请选择有效的Excel文件。")
            return
        
        # 读取 Excel 文件
        excel_file = pd.ExcelFile(file_path)
        df = excel_file.parse('Sheet1')
        df_sheet2 = excel_file.parse('Sheet2')
        
        # 进行 Vlookup 操作
        df = perform_vlookup(df, df_sheet2)
        
        # 处理数据
        df = process_data(df, cishu_value, juli_value, bili_value)
        
        # 统计结果
        low_success_count, long_distance_count = count_results(df)
        
        # 保存数据并填充颜色
        save_and_fill(df, file_path)
        
        message = f"低切换成功率个数: {low_success_count}\n远距离切换个数: {long_distance_count}"
        messagebox.showinfo("完成", f"程序执行完成！感谢使用！\n{message}")
    except ValueError:
        messagebox.showerror("错误", "请输入有效的整数！")
    except Exception as e:
        messagebox.showerror("错误", f"发生错误: {str(e)}")

# 检查输入是否完整
def check_inputs(*args):
    cishu = cishu_entry.get()
    juli = juli_entry.get()
    bili = bili_entry.get()
    file = file_entry.get()
    if cishu and juli and bili and file:
        run_button.config(state=tk.NORMAL)
    else:
        run_button.config(state=tk.DISABLED)

# 创建主窗口
root = tk.Tk()
root.title("欢迎使用：不合理切换识别工具")

# 创建文件选择部分
file_frame = tk.Frame(root)
file_frame.pack(fill=tk.X, padx=10, pady=5)

file_label = tk.Label(file_frame, text="Excel文件路径:")
file_label.pack(side=tk.LEFT)

file_entry = tk.Entry(file_frame, width=40)
file_entry.pack(side=tk.LEFT, padx=5)

browse_button = tk.Button(file_frame, text="浏览...", command=select_file)
browse_button.pack(side=tk.LEFT)

# 创建标签和输入框
cishu_label = tk.Label(root, text="请输入最小切换次数（次）")
cishu_label.pack(pady=5)
cishu_entry = tk.Entry(root, width=50, justify='center')
cishu_entry.pack()
cishu_entry.bind("<KeyRelease>", check_inputs)

juli_label = tk.Label(root, text="请输入超远切换距离（米）")
juli_label.pack(pady=5)
juli_entry = tk.Entry(root, width=50, justify='center')
juli_entry.pack()
juli_entry.bind("<KeyRelease>", check_inputs)

bili_label = tk.Label(root, text="请输入低切换成功率比例（%）")
bili_label.pack(pady=5)
bili_entry = tk.Entry(root, width=50, justify='center')
bili_entry.pack()
bili_entry.bind("<KeyRelease>", check_inputs)

# 创建运行按钮
run_button = tk.Button(root, text="运行", command=run_program, state=tk.DISABLED)
run_button.pack(pady=20)

# 添加作者信息
author_label = tk.Label(root, text="版本 V13| 作者: 小小老师")
author_label.pack(side=tk.BOTTOM, pady=7)

# 运行主循环
root.mainloop()